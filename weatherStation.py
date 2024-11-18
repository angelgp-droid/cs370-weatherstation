import smbus2
import bme280
import time
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO

# Sensor setup
address = 0x76
bus = smbus2.SMBus(1)
calibration_params = bme280.load_calibration_params(bus, address)

# Directory for saving the file
output_dir = Path.home() / "pi4Share"
output_dir.mkdir(parents=True, exist_ok=True)

# Excel file path
excel_file = output_dir / "piRecordings.xlsx"

# Initialize data lists
data_records = {
    "Time": [],
    "Temperature (°C)": [],
    "Humidity (%)": [],
    "Pressure (hPa)": []
}

# Generate scatter plots with average lines and trend lines
def create_scatter_plots(df):
    fig, axs = plt.subplots(3, 1, figsize=(10, 10))

    # Temperature scatter plot with average line and trend line
    axs[0].scatter(df['Time'], df['Temperature (°C)'], color='r')
    avg_temp = df['Temperature (°C)'].mean()
    axs[0].axhline(avg_temp, color='r', linestyle='--', label=f'Avg: {avg_temp:.2f} °C')

    # Trend line calculation for Temperature
    x_temp = np.arange(len(df['Time']))
    y_temp = df['Temperature (°C)']
    if len(df['Time']) > 1 and not y_temp.isnull().any():
        coeffs_temp = np.polyfit(x_temp, y_temp, 1)
        trend_temp = np.polyval(coeffs_temp, x_temp)
        axs[0].plot(df['Time'], trend_temp, color='darkred', linestyle='-', linewidth=2, label='Trend Line')

    axs[0].set_title('Temperature Over Time')
    axs[0].set_xlabel('Time')
    axs[0].set_ylabel('Temperature (°C)')
    axs[0].legend()

    # Humidity scatter plot with average line and trend line
    axs[1].scatter(df['Time'], df['Humidity (%)'], color='b')
    avg_humidity = df['Humidity (%)'].mean()
    axs[1].axhline(avg_humidity, color='b', linestyle='--', label=f'Avg: {avg_humidity:.2f} %')

    # Trend line calculation for Humidity
    y_humidity = df['Humidity (%)']
    if len(df['Time']) > 1 and not y_humidity.isnull().any():
        coeffs_humidity = np.polyfit(x_temp, y_humidity, 1)
        trend_humidity = np.polyval(coeffs_humidity, x_temp)
        axs[1].plot(df['Time'], trend_humidity, color='darkblue', linestyle='-', linewidth=2, label='Trend Line')

    axs[1].set_title('Humidity Over Time')
    axs[1].set_xlabel('Time')
    axs[1].set_ylabel('Humidity (%)')
    axs[1].legend()

    # Pressure scatter plot with average line and trend line
    axs[2].scatter(df['Time'], df['Pressure (hPa)'], color='g')
    avg_pressure = df['Pressure (hPa)'].mean()
    axs[2].axhline(avg_pressure, color='g', linestyle='--', label=f'Avg: {avg_pressure:.2f} hPa')

    # Trend line calculation for Pressure
    y_pressure = df['Pressure (hPa)']
    if len(df['Time']) > 1 and not y_pressure.isnull().any():
        coeffs_pressure = np.polyfit(x_temp, y_pressure, 1)
        trend_pressure = np.polyval(coeffs_pressure, x_temp)
        axs[2].plot(df['Time'], trend_pressure, color='darkgreen', linestyle='-', linewidth=2, label='Trend Line')

    axs[2].set_title('Pressure Over Time')
    axs[2].set_xlabel('Time')
    axs[2].set_ylabel('Pressure (hPa)')
    axs[2].legend()

    fig.tight_layout()
    return fig

# Function to append data to Excel and embed the scatter plots
def append_to_excel(df):
    if excel_file.exists():
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    else:
        # Create a new workbook and save data
        workbook = pd.ExcelWriter(excel_file, engine='openpyxl', mode='w')
        df.to_excel(workbook, index=False, sheet_name='SensorData')
        workbook.save()
        workbook.close()
        return

    # Get the starting row for appending
    start_row = sheet.max_row + 1
    
    # Append the data to the next available row
    for col in range(len(df.columns)):
        sheet.cell(row=start_row, column=col + 1, value=df.iloc[-1, col])  # Append only the most recent data row

    # Remove the previous scatter plot image if it exists
    for image in sheet._images:
        sheet._images.remove(image)

    # Create the scatter plot
    fig = create_scatter_plots(df)
    # Embed it in the Excel file
    image_stream = BytesIO()
    plt.savefig(image_stream, format='png')
    plt.close(fig) 
    image_stream.seek(0)  # Reset stream position
    
    img = Image(image_stream)
    img.anchor = f'E{start_row}'
    sheet.add_image(img)

    workbook.save(excel_file)

# Record data loop
record_interval = timedelta(seconds=30) 
last_record_time = datetime.now()

try:
    print("Recording sensor data every 30 seconds. Press Ctrl+C to stop.")
    while True:
        current_time = datetime.now()

        if current_time >= last_record_time + record_interval:
            data = bme280.sample(bus, address, calibration_params)

            # Extract temperature, humidity, and pressure
            temperature_celsius = data.temperature
            humidity = data.humidity
            pressure = data.pressure
            timestamp = pd.Timestamp.now()

            # Append data to lists
            data_records["Time"].append(timestamp)
            data_records["Temperature (°C)"].append(temperature_celsius)
            data_records["Humidity (%)"].append(humidity)
            data_records["Pressure (hPa)"].append(pressure)

            # Output the readings to the terminal
            print(f"{timestamp}: Temperature: {temperature_celsius:.2f} °C, "
                f"Humidity: {humidity:.2f} %, Pressure: {pressure:.2f} hPa")

            df = pd.DataFrame(data_records)

            append_to_excel(df)

            # Update the last record time
            last_record_time = current_time

        # Sleep briefly to save CPU and lower power usage
        time.sleep(55)

except KeyboardInterrupt:
    print('Program stopped')

    df = pd.DataFrame(data_records)

    append_to_excel(df)
    print(f"Data saved to {excel_file}")

except Exception as e:
    print('An unexpected error occurred:', str(e))
